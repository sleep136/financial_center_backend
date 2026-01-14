from fastapi import APIRouter, HTTPException, status
from service.program import get_one_program, get_program_info_list, get_freeze_detail, get_reimbursement_detail, \
    get_labor_cost_detail, Reimbursement, LaborCost, VoucherDetail, get_economic_classification_cost
from fastapi.responses import JSONResponse
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
import os
import tempfile

router = APIRouter()

ACCESS_TOKEN_EXPIRE_MINUTES = 30


@router.get("/program/one_program")
async def get_program(program_id: str, department_id: str):
    programs = get_one_program(program_id, department_id)
    if not programs:
        return HTTPException(status_code=status.HTTP_404_NOT_FOUND)
    return JSONResponse(status_code=200,
                        content={"data": programs})


@router.get("/program/batch_program")
async def get_batch_program(program_id: str):
    program = get_program_info_list(program_id)
    if not program:
        return HTTPException(status_code=status.HTTP_404_NOT_FOUND)
    return JSONResponse(status_code=200,
                        content={"data": program})


@router.get("/program/freeze")
async def get_batch_freeze_detail(program_id: int, department_id: int):
    freeze_details = get_freeze_detail(program_id, department_id)
    if not freeze_details:
        return HTTPException(status_code=status.HTTP_404_NOT_FOUND)
    return freeze_details


@router.get("/program/reimbursement", response_model=list[Reimbursement])
async def get_batch_reimbursement_detail(program_id: str, department_id: str, filter_state: int = 1):
    reimbursement_details = get_reimbursement_detail(program_id, department_id, filter_state)
    if not reimbursement_details:
        return HTTPException(status_code=status.HTTP_404_NOT_FOUND)
    return reimbursement_details


@router.get("/program/labor_cost", response_model=list[LaborCost])
async def get_batch_labor_cost_detail(program_id: str, department_id: str, filter_state: int = 1):
    labor_cost_details = get_labor_cost_detail(program_id, department_id, filter_state)
    if not labor_cost_details:
        return []
    return labor_cost_details


@router.get("/program/economic_classification_cost", response_model=list[VoucherDetail])
async def get_economic_classification(program_id: str, department_id: str, subject_code: str, year: int = 2026):
    economic_classification_details = get_economic_classification_cost(program_id, department_id, subject_code, year)
    if not economic_classification_details:
        return []
    return economic_classification_details


@router.get("/program/economic_classification_cost/export")
async def export_economic_classification(program_id: str, department_id: str, subject_code: str, year: int = 2026):
    """
    高级导出功能，支持查询参数
    """
    economic_classification_details = get_economic_classification_cost(program_id, department_id, subject_code, year)
    if not economic_classification_details:
        return

    # 创建临时文件
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        filename = tmp.name

        # 创建 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "导出数据"

        # 定义样式
        header_font = Font(bold=True, size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 写入表头
        headers = ["凭证号", "项目编号", "部门编号", "摘要", "科目编号", "金额"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # 写入数据
        for row, item in enumerate(economic_classification_details):

            ws.cell(row=row, column=1, value=item.voucher_number)
            ws.cell(row=row, column=2, value=item.program_id)
            ws.cell(row=row, column=3, value=item.department_id)
            ws.cell(row=row, column=4, value=item.abstract)
            ws.cell(row=row, column=5, value=item.subject_code)
            ws.cell(row=row, column=6, value=item.amount)

            # 应用边框
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = border

        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # 保存文件
        wb.save(filename)

    # 生成文件名
    export_filename = f"科研项目数据_导出_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return FileResponse(
        filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=export_filename,
        background=lambda: os.unlink(filename)  # 下载后删除临时文件
    )
